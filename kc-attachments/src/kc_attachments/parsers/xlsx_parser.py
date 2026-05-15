from __future__ import annotations
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from . import ParseResult, REGISTRY


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in row)


class XlsxParser:
    """One markdown table per sheet; formula cells emit evaluated values."""

    def parse(self, source: Path, meta: dict[str, Any]) -> ParseResult:
        # data_only=True returns the last-cached evaluated value, not the formula text.
        wb = load_workbook(str(source), data_only=True)
        parts: list[str] = []
        sheet_names: list[str] = []
        for ws in wb.worksheets:
            sheet_names.append(ws.title)
            parts.append(f"## {ws.title}")
            rows = [tuple(c.value for c in r) for r in ws.iter_rows()]
            while rows and _is_empty_row(rows[-1]):
                rows.pop()
            if not rows:
                parts.append("_(empty sheet)_")
                continue
            header = rows[0]
            header_cells = ["" if v is None else str(v) for v in header]
            sep_cells = ["---" for _ in header_cells]
            lines = [
                "| " + " | ".join(header_cells) + " |",
                "| " + " | ".join(sep_cells) + " |",
            ]
            for r in rows[1:]:
                cells = ["" if v is None else str(v) for v in r]
                lines.append("| " + " | ".join(cells) + " |")
            parts.append("\n".join(lines))
        return ParseResult(
            markdown="\n\n".join(parts),
            extra_meta={"sheet_names": sheet_names},
        )


REGISTRY["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"] = XlsxParser()
